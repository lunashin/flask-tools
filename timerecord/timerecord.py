#!/usr/bin/python3

###############################################
# 工数入力情報作成
#
# 手順：
#  1.GoogleタイムカードからCSVをエクスポート (YYYYMM-user_time_card.csv)
#  2.DDS勤怠のエクセルから該当範囲をコピーし、テキストファイルへ保存 (YYYYMM-timecard.csv)
#  3.上記２ファイルを本スクリプトと同じディレクトリに配置する。
#  3.本スクリプトを実行
#
###############################################

import openpyxl as px
import datetime
import csv
import calendar
import sys
import os




################################################
# Class
################################################

# レコード管理
class TimeRecords:
    records = None      # TimeRecord[]
    xls_sheet_name = "" # xlsの参照シート名
    adjust_hour = 1    # 時間調整(昼休憩など)

    # コンストラクタ
    # xls_sheet_name: xlsのシート名
    def __init__(self, xls_sheet_name) -> None:
        self.records = []
        self.xls_sheet_name = xls_sheet_name

    # パース (ファイル一覧内の最初のcsv,xlsxファイルをパース)
    # file_list: ファイル一覧
    # return True/False
    def parse(self, file_list):
        csv_file = ""
        xlsx_file = ""

        # 拡張子からファイル種類を判定
        for file in file_list:
            root, ext = os.path.splitext(file)
            if ext == '.csv' and csv_file == '':
                csv_file = file
            if ext == '.xlsx' and xlsx_file == '':
                xlsx_file = file
        
        # どちらかのファイルが見つからない
        if csv_file == '' or xlsx_file == '':
            return False
        
        # ファイル存在確認
        if not os.path.isfile(csv_file):
            return False
        if not os.path.isfile(xlsx_file):
            return False
        
        # パース
        self.parse_from_google_csv(csv_file)
        self.parse_from_xls(xlsx_file)

        return True

    # Google勤怠からパース
    def parse_from_google_csv(self, file_csv):
        # google timecart csv から TimeRecord クラスを生成し、リストへ追加
        with open(file_csv, encoding="shift_jis") as f:
            r = csv.reader(f)
            for row in r:
                if row[2] == "date_stamp":
                    continue
                if row[3] != "":
                    # print(row[2] + "\t" + row[3] + "\t" + row[4])
                    start = datetime.datetime.strptime(row[2] + " " + row[3], '%Y-%m-%d %H:%M')
                    end = datetime.datetime.strptime(row[2] + " " + row[4], '%Y-%m-%d %H:%M')
                    obj = TimeRecord(start, end, True)
                    self.records.append(obj)
        # ソート
        self.records = sorted(self.records, key=lambda obj: obj.start)

    # xlsファイルからパース
    def parse_from_xls(self, file_xml):
        wb = px.load_workbook(file_xml)
        # select sheet
        print("- select sheet")
        ws = wb[self.xls_sheet_name]
        # read from xls
        print("- read data")
        for num in range(1, 31):
            day = ws.cell(column=2, row=num+2).value
            start = ws.cell(column=3, row=num+2).value
            end = ws.cell(column=4, row=num+2).value
            if day != None:
                if start != None and end != None:
                    start_time = day + start
                    end_time = day + end
                    obj = TimeRecord(start_time, end_time, False)
                    self.records.append(obj)
        # ソート
        self.records = sorted(self.records, key=lambda obj: obj.start)

    # 勤務時間のみのCSV文字列を取得
    def get_csv_worktime(self):
        lines = []
        for item in self.records:
            worktime = item.get_worktime(self.adjust_hour)
            lines.append('{0:.2f}'.format(worktime))
        return '\n'.join(lines)

    # CSV文字列を取得
    def get_csv(self, delimiter, only_worktime=False):
        year = self.records[0].start.year
        month = self.records[0].start.month
        d = datetime.date( year , month , 1 )

        lines = []
        
        # 1日 - 月末の日 まで回す
        for i in range(calendar.monthrange(year, month)[1]):
            hit = False
            for item in self.records:
                if item.start.date() == d:
                    if only_worktime:
                        worktime = item.get_worktime(self.adjust_hour)
                        lines.append('{0:.2f}'.format(worktime))
                    else:
                        lines.append(item.getLine(delimiter))
                    hit = True
                    break
            # リスト中にデータがなければ、空行を出力する
            if hit == False:
                s = ''
                if only_worktime == False:
                    s = d.strftime('%Y/%m/%d') + "\t" + "\t"
                lines.append(s)

            d += datetime.timedelta(days=1)

        # 結果を文字列化して返す
        print(lines)
        return '\n'.join(lines)



# レコード
class TimeRecord:
    start = datetime.datetime.now() # 開始日時
    end = datetime.datetime.now()   # 終了日時
    remote = False                  # リモートかどうか

    def __init__(self, start_time, end_time, remote):
        self.start = start_time
        self.end = end_time
        self.remote = remote

    # 勤務時間取得 (0.25単位)
    def get_worktime(self, adjust_hour):
        dt = self.end - self.start
        hour = (dt.seconds / 60 / 60)

        # 調整(昼休憩などの時間)
        hour = hour - adjust_hour
        
        # 0.25単位に丸める
        return self.rounding(hour)

    # CSV行取得
    def getLine(self, delimiter):
        remote_str = ""
        if self.remote == True:
            remote_str = "Remote"

        line = []
        line.append(self.start.strftime('%Y/%m/%d'))
        line.append(self.start.strftime('%H:%M:%S'))
        line.append(self.end.strftime('%H:%M:%S'))
        line.append(remote_str)
        return delimiter.join(line)

    # 丸め (0.25単位)
    def rounding(self, value):
        after_point = value - int(value)
        if after_point < 0.25:
            after_point = 0
        elif after_point < 0.5:
            after_point = 0.25
        elif after_point < 0.75:
            after_point = 0.5
        elif after_point >= 0.75:
            after_point = 0.75
        return int(value) + after_point
